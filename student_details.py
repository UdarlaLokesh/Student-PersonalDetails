import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook

root = tk.Tk()
root.title("Student details")

buttons_frame = ttk.LabelFrame(root, text=' Student Details ')
buttons_frame.grid(column=0, row=1)

ttk.Label(buttons_frame, text="Enter Student's Name:").grid(column=0, row=0)
ttk.Label(buttons_frame, text="Enter Student's Age:").grid(column=0, row=1)
ttk.Label(buttons_frame, text="Enter Student's Roll Number:").grid(column=0, row=2)
ttk.Label(buttons_frame, text="Enter Student's Phone Number:").grid(column=0, row=3)
ttk.Label(buttons_frame, text="Enter Student's Marks:").grid(column=0, row=4)

name = tk.StringVar()
name_entered = ttk.Entry(buttons_frame, width=24, textvariable=name)
name_entered.grid(column=1, row=0)

roll_number = tk.StringVar()
roll_number_entered = ttk.Entry(buttons_frame, width=24, textvariable=roll_number)
roll_number_entered.grid(column=1, row=2)

age = tk.StringVar()
age_entered = ttk.Entry(buttons_frame, width=24, textvariable=age)
age_entered.grid(column=1, row=1)

phone_number = tk.StringVar()
phone_number_entered = ttk.Entry(buttons_frame, width=24, textvariable=phone_number)
phone_number_entered.grid(column=1, row=3)

marks = tk.StringVar()
marks_entered = ttk.Entry(buttons_frame, width=24, textvariable=marks)
marks_entered.grid(column=1, row=4)

table_frame = ttk.LabelFrame(root, text=' Table ')
table_frame.grid(column=0, row=8)

tv = ttk.Treeview(table_frame, columns=(1, 2, 3, 4, 5), show="headings", height="5")
tv.pack()
tv.heading(1, text="Student's Name")
tv.heading(2, text="Student's Roll Number")
tv.heading(3, text="Student's Age")
tv.heading(4, text="Student's Phone Number")
tv.heading(5, text="Student's Marks")

wb = Workbook()
ws = wb.active
ws.append(["Student's Name", "Student's Roll Number", "Student's Age", "Student's Phone Number", "Student's Marks"])  # header row

def action():
    tv.insert('', 'end', values=(name.get(), roll_number.get(), age.get(), phone_number.get(), marks.get()))
    ws.append([name.get(), roll_number.get(), age.get(), phone_number.get(), marks.get()])  # add row to Excel file
    name_entered.delete(0, 'end')
    roll_number_entered.delete(0, 'end')
    age_entered.delete(0, 'end')
    phone_number_entered.delete(0, 'end')
    marks_entered.delete(0, 'end')

action_button = ttk.Button(buttons_frame, text="Insert Details", command=action)
action_button.grid(column=2, row=0)

def delete_row():
    selected_item = tv.selection()[0]
    tv.delete(selected_item)
    row_index = tv.index(selected_item) + 2  # adjust for header row
    ws.delete_rows(row_index)  # delete row from Excel file

delete_button = ttk.Button(buttons_frame, text="Delete Row", command=delete_row)
delete_button.grid(column=2, row=2)

def update_row():
    selected_item = tv.selection()[0]
    tv.item(selected_item, values=(name.get(), roll_number.get(), age.get(), phone_number.get(), marks.get()))
    row_index = tv.index(selected_item) + 2  # adjust for header row
    ws.cell(row=row_index, column=1).value = name.get()
    ws.cell(row=row_index, column=2).value = roll_number.get()
    ws.cell(row=row_index, column=3).value = age.get()
    ws.cell(row=row_index, column=4).value = phone_number.get()
    ws.cell(row=row_index, column=5).value = marks.get()

update_button = ttk.Button(buttons_frame, text="Update Row", command=update_row)
update_button.grid(column=2, row=3)

def save_excel_file():
    wb.save("student_details.xlsx")  # save Excel file

save_button = ttk.Button(buttons_frame, text="Save to Excel", command=save_excel_file)
save_button.grid(column=2, row=1)

root.mainloop()
